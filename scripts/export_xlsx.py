#!/usr/bin/env python3
"""Export case-gen structured cases to an Excel workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any


HEADERS = [
    "Case ID",
    "Title",
    "Priority",
    "Level",
    "Type",
    "Suite Layers",
    "Traceability",
    "Preconditions",
    "Data",
    "Steps",
    "Expected",
    "Automation",
    "Status",
]


def load_data(path: Path) -> Any:
    text = path.read_text(encoding="utf-8")
    if path.suffix.lower() in {".yaml", ".yml"}:
        try:
            import yaml  # type: ignore
        except ImportError as exc:
            raise RuntimeError("YAML input requires PyYAML. Use JSON or install PyYAML.") from exc
        return yaml.safe_load(text)
    return json.loads(text)


def as_list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def flatten(value: Any) -> str:
    if isinstance(value, list):
        return "\n".join(flatten(item) for item in value if flatten(item))
    if isinstance(value, dict):
        return "\n".join(f"{key}: {flatten(val)}" for key, val in value.items() if flatten(val))
    return str(value or "")


def trace_summary(case: dict[str, Any]) -> str:
    trace = case.get("traceability")
    if not isinstance(trace, dict):
        return ""
    lines: list[str] = []
    for key in ("requirements", "code", "api", "rules", "risks", "test_points"):
        values = as_list(trace.get(key))
        if values:
            lines.append(f"{key}: {', '.join(str(v) for v in values)}")
    if trace.get("inferred"):
        lines.append("inferred: true")
    return "\n".join(lines)


def suite_layers_for(case_id: str, data: dict[str, Any]) -> str:
    layers = data.get("suite_layers")
    if not isinstance(layers, dict):
        return ""
    return ", ".join(name for name, ids in layers.items() if case_id in [str(i) for i in as_list(ids)])


def expected_summary(case: dict[str, Any]) -> str:
    lines: list[str] = []
    for item in as_list(case.get("expected")):
        if isinstance(item, dict):
            observable = item.get("observable", "")
            assertion = item.get("assertion", "")
            lines.append(f"[{observable}] {assertion}")
        else:
            lines.append(str(item))
    return "\n".join(lines)


def export(data: dict[str, Any], output: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("export_xlsx.py requires openpyxl. Install it with: pip install openpyxl") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Cases"

    for column, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="366092")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row, case in enumerate(as_list(data.get("cases")), start=2):
        if not isinstance(case, dict):
            continue
        case_id = str(case.get("id") or "")
        values = [
            case_id,
            case.get("title"),
            case.get("priority"),
            case.get("level"),
            case.get("type"),
            suite_layers_for(case_id, data),
            trace_summary(case),
            flatten(case.get("preconditions")),
            flatten(case.get("data")),
            flatten(case.get("steps")),
            expected_summary(case),
            case.get("automation_candidate"),
            case.get("status"),
        ]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [18, 42, 10, 14, 18, 24, 42, 42, 36, 52, 58, 14, 12]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width

    ws.freeze_panes = "A2"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export case-gen cases to xlsx.")
    parser.add_argument("input", type=Path)
    parser.add_argument("-o", "--output", type=Path, required=True)
    args = parser.parse_args()

    data = load_data(args.input)
    if not isinstance(data, dict):
        raise RuntimeError("Top-level document must be an object.")
    export(data, args.output)
    print(f"Exported xlsx: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

